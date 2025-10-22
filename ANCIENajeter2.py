import streamlit as st
import pandas as pd
import re
from typing import Optional, List, Dict, Tuple
from io import BytesIO
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import itertools

st.set_page_config(page_title="Portefeuille Logiciels – Consolidation & Export PI-Planner", layout="wide")
st.title("Portefeuille multi-centres")
st.title("Jules Devos V.1.0.3")


# ===================== Helpers génériques =====================

def norm(s: str) -> str:
    """Minuscule + supprime espaces/ponctuation (pour comparaisons robustes)."""
    if s is None:
        return ""
    x = str(s).strip().lower()
    x = re.sub(r"[^a-z0-9]+", "", x)
    return x

def find_col(df: pd.DataFrame, aliases: List[str]) -> Optional[str]:
    """Trouve une colonne en testant une liste d'alias (normalisés)."""
    norm_cols = {norm(c): c for c in df.columns}
    for alias in aliases:
        a = norm(alias)
        if a in norm_cols:
            return norm_cols[a]
    return None

def get_bytes(file) -> Tuple[str, bytes]:
    """Retourne (nom, bytes) d'un UploadedFile (ou file-like)."""
    name = getattr(file, "name", "uploaded")
    try:
        file.seek(0)
    except Exception:
        pass
    data = file.read()
    return name, data

def parse_csv_or_excel(file):
    """Charge CSV (essaye ; puis ,) ou Excel, en lisant d'abord en mémoire (BytesIO)."""
    if file is None:
        return None
    name, data = get_bytes(file)
    bio = BytesIO(data)
    if name.lower().endswith((".xlsx", ".xls")):
        return pd.read_excel(bio)
    # CSV
    try:
        bio.seek(0)
        return pd.read_csv(bio, delimiter=";", quotechar='"')
    except Exception:
        bio.seek(0)
        return pd.read_csv(bio)

def read_excel_header3_from_bytes(data: bytes) -> pd.DataFrame:
    """Lit un Excel en mémoire avec header=3 (4e ligne = en-têtes)."""
    return pd.read_excel(BytesIO(data), header=3)

# ===================== Centres (mapping + détection) =====================

# Mapping canon
CENTER_CANON = {
    "LNE": "Lille",
    "NGE": "Nancy",
    "SIF": "Saclay",
    "PRO": "Paris",
    "RBA": "Rennes",
    "SAM": "Sophia",
    "BSO": "Bordeaux",
    "GRA": "Grenoble",
    "LYS": "Lyon",
}

# Synonymes/indices pour auto-détection (noms de fichiers/onglets/1res lignes)
CENTER_SYNONYMS = {
    "lne": ("LNE", "Lille"),
    "lille": ("LNE", "Lille"),
    "nge": ("NGE", "Nancy"),
    "nancy": ("NGE", "Nancy"),
    "sif": ("SIF", "Saclay"),
    "saclay": ("SIF", "Saclay"),
    "idf": ("SIF", "Saclay"),
    "iledefrance": ("SIF", "Saclay"),
    "SAC": ("SIF", "Saclay"),
    "pro": ("PRO", "Paris"),
    "paris": ("PRO", "Paris"),
    "rocquencourt": ("PRO", "Paris"),
    "rba": ("RBA", "Rennes"),
    "rennes": ("RBA", "Rennes"),
    "sam": ("SAM", "Sophia"),
    "sophia": ("SAM", "Sophia"),
    "sophiaantipolis": ("SAM", "Sophia"),
    "bso": ("BSO", "Bordeaux"),
    "bordeaux": ("BSO", "Bordeaux"),
    "gra": ("GRA", "Grenoble"),
    "grenoble": ("GRA", "Grenoble"),
    "uga": ("GRA", "Grenoble"),
    "lys": ("LYS", "Lyon"),
    "lyon": ("LYS", "Lyon"),
}

def detect_center_from_text(text: str) -> Tuple[Optional[str], Optional[str]]:
    t = norm(text)
    # Codes canons
    for code, ville in CENTER_CANON.items():
        if norm(code) in t:
            return code, ville
    # Synonymes étendus
    for key, (code, ville) in CENTER_SYNONYMS.items():
        if key in t:
            return code, ville
    return None, None

def auto_detect_file_center(file_name: str, file_bytes: bytes) -> Tuple[str, str]:
    # Devine le centre à partir du nom de fichier, des noms d'onglets et du contenu des 5 premières lignes de chaque onglet. Fallback: AUTRES.

    # 1) Nom de fichier
    code, ville = detect_center_from_text(file_name)
    if code:
        return code, ville
    # 2) Noms d’onglets
    try:
        xls = pd.ExcelFile(BytesIO(file_bytes))
        for sh in xls.sheet_names:
            code, ville = detect_center_from_text(sh)
            if code:
                return code, ville
        # 3) 5 premières lignes par onglet
        for sh in xls.sheet_names:
            head = pd.read_excel(BytesIO(file_bytes), sheet_name=sh, header=None, nrows=5)
            flat = " ".join([str(v) for v in itertools.chain.from_iterable(head.values.tolist()) if pd.notna(v)])
            code, ville = detect_center_from_text(flat)
            if code:
                return code, ville
    except Exception:
        pass
    return "AUTRES", "Autres"

# ===================== Modèle de colonnes "canon" =====================

CANON = [
    "Nom du logiciel", "Référence BIL", "Référence contrat (legisway)",
    "Centres Inria impliqués", "Type de licence logiciel",
    "Valorisation", "Description (BIL)", "Date de dépôt APP", "IDDN",
    "Équipe", "Auteurs et parts", "Logo ?", "Commentaires"
]

ALIASES: Dict[str, List[str]] = {
    "Nom du logiciel": ["Nom du logiciel", "Logiciel", "Libellé", "Nom"],
    "Référence BIL": ["Référence BIL", "Ref BIL", "BIL", "Réf BIL"],
    "Référence contrat (legisway)": ["Référence contrat (legisway)", "Référence contrat", "Reference contrat", "Ref contrat", "Contrat", "Legisway", "N° contrat", "Numéro de contrat", "Numero de contrat", "Contract number", "Contract ID", "ID contrat"],
    "Centres Inria impliqués": ["Centres Inria impliqués", "Centre Inria", "Centre", "Centre déposant", "Centre deposant"],
    "Type de licence logiciel": ["Type de licence logiciel", "Licence", "Type licence", "Mots clés"],
    "Valorisation": ["Valorisation (licence, cession, projet, dormant, consortium)", "Valorisation"],
    "Description (BIL)": ["Description (BIL)", "Description", "Desc BIL"],
    "Date de dépôt APP": ["Date de dépôt APP", "Date de dépôt", "Date depot", "Date APP"],
    "IDDN": ["IDDN", "Num IDDN"],
    "Équipe": ["Equipe du projet", "Équipe du projet", "Équipe", "Equipe"],
    "Auteurs et parts": ["Auteurs et leurs parts", "Auteurs", "Auteurs et parts"],
    "Logo ?": ["Est-ce qu'il y a un logo ?", "Logo ?", "Logo"],
    "Commentaires": ["Commentaires", "Notes", "Remarques"]
}

def unify_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame()
    for canon in CANON:
        col = find_col(df, ALIASES[canon])
        out[canon] = df[col] if col else ""
    out["__nom_lower"] = out["Nom du logiciel"].astype(str).str.strip().str.lower()
    out["__bil_num"] = out["Référence BIL"].astype(str).str.extract(r"(\d+)")
    return out

# ===================== Legisway (contrat) =====================

LEGIS_NUM_ALIASES = [
    "Numéro de contrat", "Numero de contrat", "N° contrat", "Contract number",
    "Contract ID", "ID contrat", "Référence contrat", "Reference contrat", "Réf contrat", "Ref contrat"
]
LEGIS_NAME_ALIASES = ["Nom du logiciel", "Logiciel", "Famille", "Libellé", "Nom"]

def build_legis_index(legis_df: Optional[pd.DataFrame]) -> Tuple[Optional[pd.DataFrame], List[str], Optional[str]]:
    if legis_df is None:
        return None, [], None
    legis_num_cols = []
    for alias in LEGIS_NUM_ALIASES:
        c = find_col(legis_df, [alias])
        if c and c not in legis_num_cols:
            legis_num_cols.append(c)
    legis_name_col = find_col(legis_df, LEGIS_NAME_ALIASES)
    if legis_name_col:
        legis_df["__legis_name_lower"] = legis_df[legis_name_col].astype(str).str.strip().str.lower()
    else:
        legis_df["__legis_name_lower"] = ""
    if legis_num_cols:
        has_num = pd.Series(False, index=legis_df.index)
        for c in legis_num_cols:
            has_num = has_num | legis_df[c].astype(str).str.strip().str.contains(r"\d")
        legis_df["__legis_has_number"] = has_num
    else:
        legis_df["__legis_has_number"] = False
    return legis_df, legis_num_cols, legis_name_col

def has_contract_number(row: pd.Series, legis_df: Optional[pd.DataFrame], legis_name_col: Optional[str]) -> bool:
    # 1) Référence contrat présente dans le portefeuille (au moins un chiffre)
    ref_col = "Référence contrat (legisway)"
    ref_val = str(row.get(ref_col, "")).strip()
    if ref_val and re.search(r"\d", ref_val):
        return True
    # 2) Sinon, match nom logiciel avec Legisway et vérifier un numéro
    if legis_df is not None and legis_name_col is not None:
        nom = str(row.get("Nom du logiciel", "")).strip().lower()
        if not nom:
            return False
        subset = legis_df[legis_df["__legis_name_lower"] == nom]
        if not subset.empty and subset["__legis_has_number"].any():
            return True
    return False

def has_keyword_valo(row: pd.Series, keywords: List[str]) -> bool:
    """Protégé par une licence si le champ 'Type de licence logiciel' contient un mot-clé choisi."""
    val = str(row.get("Type de licence logiciel", "")).strip().lower()
    if not val:
        return False
    for kw in keywords:
        if kw.lower() in val:
            return True
    return False

# ===================== BIL (lookup & "Mise à jour" si existant) =====================

BIL_NAME_ALIASES = ["Logiciel", "Nom du logiciel", "Libellé", "Nom"]
BIL_DATE_ALIASES = ["Date de dépôt", "Date depot", "Date_dépôt", "Date APP"]
BIL_IDDN_ALIASES = ["Num IDDN", "IDDN"]
BIL_CENTRE_ALIASES = ["Centre déposant", "Centre deposant", "Centre", "Centre Inria"]

def build_bil_lookup(df_bil: Optional[pd.DataFrame]) -> Tuple[Optional[pd.DataFrame], Optional[str], Optional[str], Optional[str]]:
    if df_bil is None:
        return None, None, None, None
    name_col = find_col(df_bil, BIL_NAME_ALIASES)
    date_col = find_col(df_bil, BIL_DATE_ALIASES)
    iddn_col = find_col(df_bil, BIL_IDDN_ALIASES)
    centre_col = find_col(df_bil, BIL_CENTRE_ALIASES)
    if name_col:
        df_bil["__bil_name_lower"] = df_bil[name_col].astype(str).str.strip().str.lower()
    else:
        df_bil["__bil_name_lower"] = ""
    return df_bil, date_col, iddn_col, centre_col

def compute_mise_a_jour(row: pd.Series, bil_df: Optional[pd.DataFrame], bil_date_col: Optional[str], bil_iddn_col: Optional[str], bil_centre_col: Optional[str]) -> Tuple[bool, str]:
    if bil_df is None:
        return False, ""
    name = str(row.get("Nom du logiciel", "")).strip().lower()
    if not name:
        return False, ""
    sub = bil_df[bil_df["__bil_name_lower"] == name]
    if sub.empty:
        return False, ""  # pas trouvé dans BIL → pas de diff détectable
    reasons = []
    # Date de dépôt
    port_date = str(row.get("Date de dépôt APP", "")).strip()
    bil_date = str(sub.iloc[0][bil_date_col]).strip() if bil_date_col else ""
    if bil_date and port_date and port_date != bil_date:
        reasons.append("Date de dépôt différente")
    elif bil_date and not port_date:
        reasons.append("Date de dépôt manquante (portefeuille)")
    # IDDN
    port_iddn = str(row.get("IDDN", "")).strip()
    bil_iddn = str(sub.iloc[0][bil_iddn_col]).strip() if bil_iddn_col else ""
    if bil_iddn and port_iddn and norm(port_iddn) != norm(bil_iddn):
        reasons.append("IDDN différent")
    elif bil_iddn and not port_iddn:
        reasons.append("IDDN manquant (portefeuille)")
    # Centre
    port_centre = str(row.get("Centres Inria impliqués", "")).strip()
    bil_centre = str(sub.iloc[0][bil_centre_col]).strip() if bil_centre_col else ""
    if bil_centre and port_centre and norm(port_centre) != norm(bil_centre):
        reasons.append("Centre différent")
    elif bil_centre and not port_centre:
        reasons.append("Centre manquant (portefeuille)")
    return (len(reasons) > 0, ", ".join(reasons))

# ===================== Constit de l'Excel avec 1 feuille par centre =====================

def build_excel_by_center(df_valo: pd.DataFrame) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alt_fill = PatternFill("solid", fgColor="F2F2F2")

    # Détermination centre (via colonne canon “Centres Inria impliqués” déjà normalisée)
    df_valo = df_valo.copy()

    # S’assure que la colonne est bien au format "CODE - Ville" si possible
    def normalize_center_val(s):
        code, ville = detect_center_from_text(str(s))
        if code:
            return f"{code} - {ville}"
        return str(s) if s else "AUTRES - Autres"

    df_valo["Centres Inria impliqués"] = df_valo["Centres Inria impliqués"].apply(normalize_center_val)

    # Groupement par centre
    centres_split = df_valo["Centres Inria impliqués"].fillna("AUTRES - Autres").str.split(" - ", n=1, expand=True)
    df_valo["__centre_code"] = centres_split[0]
    df_valo["__centre_ville"] = centres_split[1].fillna("Autres")

    for (code, ville), sub in df_valo.groupby(["__centre_code", "__centre_ville"], dropna=False):
        sheet_name = f"{code} - {ville}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        headers = CANON
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
            ws.column_dimensions[get_column_letter(c)].width = 28

        for i, (_, r) in enumerate(sub.iterrows(), start=2):
            ws.append([r.get(h, "") for h in headers])
            if i % 2 == 0:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=i, column=c).fill = alt_fill
        ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# ===================== Export PI-Planner =====================

def build_pi_planner_export_from_df(df_valo: pd.DataFrame) -> pd.DataFrame:
    def deposant_from_row(r):
        # "Inria Ville" si le centre est connu
        raw = str(r.get("Centres Inria impliqués", ""))
        code, ville = detect_center_from_text(raw)
        if ville:
            return f"Inria {ville}"
        return "Inria"

    def crit3_from_row(r):
        raw = str(r.get("Centres Inria impliqués", ""))
        code, ville = detect_center_from_text(raw)
        return ville if ville else ""

    out = pd.DataFrame({
        "Famille": df_valo["Nom du logiciel"],
        "Titre interne": "",
        "Déposant": df_valo.apply(deposant_from_row, axis=1),
        "Titulaire(s)": "Inria",  # ajustable si injection d'un vrai titulaire
        "Mandataire": "Inria",
        "Critère 1": "",
        "Critère 2": "",
        "Critère 3": df_valo.apply(crit3_from_row, axis=1),
        "Auteur(s)": "",
        "Responsable": "",
        "Valorisation": df_valo.get("Valorisation", ""),
        "Dépositaire": "",
        "Date de dépôt": df_valo.get("Date de dépôt APP", "")
    })
    return out

# ===================== UI Uploads =====================

st.subheader("1) Charger les fichiers")
portfolios = st.file_uploader("Portefeuilles (Excel) – sélection multiple", type=["xlsx"], accept_multiple_files=True)
file_bil = st.file_uploader("Export BIL (CSV)", type=["csv"])
file_ip = st.file_uploader("Export PI-Planner (FamillesBrevet.csv)", type=["csv"])
file_legis = st.file_uploader("Export Legisway (CSV/Excel) – optionnel", type=["csv", "xlsx", "xls"])

with st.expander("⚙️ Paramètres de filtrage"):
    kw_input = st.text_input("Mot-clé de licence à considérer comme VALORISABLE (ex: licence propriétaire)",
                             value="licence propriétaire")
    license_keywords = [k.strip() for k in kw_input.split(",") if k.strip()]

# ---- UI override centres par fichier ----
st.subheader("1bis) Centres détectés (corrige si besoin)")
override_centers: Dict[str, Tuple[str, str]] = {}
if portfolios:
    for f in portfolios:
        fname, fbytes = get_bytes(f)
        auto_code, auto_city = auto_detect_file_center(fname, fbytes)
        options = [("AUTRES", "Autres")] + list(CENTER_CANON.items())
        labels = [f"{c} – {v}" for c, v in options]
        default_idx = 0
        for i, (c, v) in enumerate(options):
            if c == auto_code:
                default_idx = i
                break
        sel = st.selectbox(
            f"Centre pour « {fname} »",
            options=list(range(len(options))),
            index=default_idx,
            format_func=lambda i: labels[i]
        )
        chosen_code, chosen_city = options[sel]
        override_centers[fname] = (chosen_code, chosen_city)

st.markdown("---")

# ===================== Pipeline principal =====================

if st.button("Consolider, filtrer et préparer l'export"):
    # ---- PI-Planner (pour info/doublons éventuels) ----
    if not portfolios:
        st.error("Merci de charger au moins un fichier **Portefeuille**.")
        st.stop()
    if not file_ip:
        st.error("Merci de charger l'**export PI-Planner** (FamillesBrevet.csv).")
        st.stop()

    df_ip = parse_csv_or_excel(file_ip)
    col_ip_reference = find_col(df_ip, ["Référence", "Reference"])
    if col_ip_reference is not None:
        df_ip["__ip_bil"] = df_ip[col_ip_reference].astype(str).str.extract(r"(\d+)")
    col_ip_famille = find_col(df_ip, ["Famille de brevet", "Famille", "Libellé", "Nom"])
    existing_bil = set(df_ip["__ip_bil"].dropna().astype(str)) if "__ip_bil" in df_ip.columns else set()
    existing_names = set(df_ip[col_ip_famille].dropna().astype(str).str.strip().str.lower()) if col_ip_famille else set()

    # ---- Legisway ----
    legis_df_raw = parse_csv_or_excel(file_legis) if file_legis else None
    legis_df, legis_num_cols, legis_name_col = build_legis_index(legis_df_raw)

    # ---- BIL pour "Mise à jour" ----
    df_bil_raw = parse_csv_or_excel(file_bil) if file_bil else None
    bil_df, bil_date_col, bil_iddn_col, bil_centre_col = build_bil_lookup(df_bil_raw)

    # ---- Unification & consolidation des portefeuilles ----
    unified_all = []
    for f in portfolios:
        fname, fbytes = get_bytes(f)
        try:
            df_raw = read_excel_header3_from_bytes(fbytes)
        except Exception:
            df_raw = pd.read_excel(BytesIO(fbytes))  # fallback
        uni = unify_columns(df_raw)

        # Si la colonne centre est vide/peu claire → injecte override sélectionné
        center_cell = str(uni["Centres Inria impliqués"].iloc[0]) if len(uni) else ""
        if not center_cell.strip():
            code, ville = override_centers.get(fname, auto_detect_file_center(fname, fbytes))
            uni["Centres Inria impliqués"] = f"{code} - {ville}"
        else:
            # même si renseigné, pose au format "CODE - Ville" si on reconnaît un synonyme
            code, ville = detect_center_from_text(center_cell)
            if code:
                uni["Centres Inria impliqués"] = f"{code} - {ville}"

        unified_all.append(uni)

    df_all = pd.concat(unified_all, ignore_index=True)

    # ---- Flags "Valorisable" & Mise à jour ----
    df_all["__has_contract_number"] = df_all.apply(lambda r: has_contract_number(r, legis_df, legis_name_col), axis=1)
    df_all["__has_kw_valo"] = df_all.apply(lambda r: has_keyword_valo(r, license_keywords), axis=1)
    df_all["Valorisable"] = df_all["__has_contract_number"] | df_all["__has_kw_valo"]

    if bil_df is not None:
        maj_flags, maj_reasons = [], []
        for _, row in df_all.iterrows():
            flag, reason = compute_mise_a_jour(row, bil_df, bil_date_col, bil_iddn_col, bil_centre_col)
            maj_flags.append(flag)
            maj_reasons.append(reason)
        df_all["Mise à jour"] = maj_flags
        df_all["Raison MAJ"] = maj_reasons
    else:
        df_all["Mise à jour"] = False
        df_all["Raison MAJ"] = ""

    # ---- Affichages ----
    st.subheader("2) TOUS les logiciels extraits")
    st.caption("Utilise les colonnes 'Valorisable' et 'Mise à jour' pour filtrer.")
    st.dataframe(df_all[CANON + ["Valorisable", "Mise à jour", "Raison MAJ"]], use_container_width=True)

    st.subheader("3) Logiciels non valorisables (seront exclus)")
    df_non_valo = df_all[~df_all["Valorisable"]]
    st.dataframe(df_non_valo[CANON], use_container_width=True)

    st.subheader("4) Logiciels valorisables (seront conservés & exportés)")
    df_valo = df_all[df_all["Valorisable"]].copy()
    st.dataframe(df_valo[CANON + ["Mise à jour", "Raison MAJ"]], use_container_width=True)

    # ---- Excel consolidé (1 feuille par centre) ----
    st.markdown("---")
    st.subheader("5) Générer l’Excel consolidé (1 feuille par centre)")
    excel_bytes = build_excel_by_center(df_valo)
    today_str = datetime.now(ZoneInfo("Europe/Paris")).strftime("%d-%m-%Y")
    excel_name = f"{today_str}-portefeuille-centres.xlsx"
    st.download_button(
        "Télécharger l’Excel consolidé",
        data=excel_bytes,
        file_name=excel_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # ---- Export PI-Planner depuis les valorisables ----
    st.subheader("6) Export PI-Planner (depuis les valorisables)")
    pi_export = build_pi_planner_export_from_df(df_valo)
    st.dataframe(pi_export.head(20), use_container_width=True)
    st.download_button(
        "Télécharger le CSV d'import PI-Planner",
        pi_export.to_csv(index=False, sep=";", encoding="utf-8").encode("utf-8"),
        file_name="import_pi_planner.csv",
        mime="text/csv"
    )

    with st.expander("🔧 Debug colonnes détectées"):
        if legis_df is not None:
            st.write("Legisway:", list(legis_df.columns))
        if df_bil_raw is not None:
            st.write("BIL:", list(df_bil_raw.columns))
        st.write("PI-Planner:", list(df_ip.columns))
        st.write("Alias portefeuille utilisés:", ALIASES)
