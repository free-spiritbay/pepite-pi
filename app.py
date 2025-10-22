import streamlit as st
import pandas as pd
import re
from typing import Optional, List, Dict, Tuple, Any
from io import BytesIO
from datetime import datetime
from zoneinfo import ZoneInfo
import itertools, os

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

# ================== En-tête ==================
import streamlit as st
from pathlib import Path

APP_VERSION = "péPIte planner V1.6 - Jules Devos"

if Path("logo.png").exists():
    st.image("logo.svg", width=250)

st.title("🍪 Générateur PI-Planner")
st.caption(APP_VERSION) 

# ================== HELPERS GÉNÉRIQUES ==================
def norm(s: Any) -> str:
    if s is None: return ""
    x = str(s).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", x)

def find_col(df: pd.DataFrame, aliases: List[str]) -> Optional[str]:
    cols = {norm(c): c for c in df.columns}
    for a in aliases:
        if norm(a) in cols:
            return cols[norm(a)]
    return None

def get_bytes(file) -> Tuple[str, bytes]:
    if file is None: return "", b""
    name = getattr(file, "name", "uploaded")
    try: file.seek(0)
    except Exception: pass
    data = file.read()
    return name, data

def parse_csv_or_excel(file) -> Optional[pd.DataFrame]:
    if file is None: return None
    name, data = get_bytes(file)
    if not data: return None
    bio = BytesIO(data)
    if name.lower().endswith((".xlsx", ".xls")):
        return pd.read_excel(bio)
    for sep in [";", ",", "\t"]:
        try:
            bio.seek(0)
            df = pd.read_csv(bio, sep=sep, engine="python", dtype=str)
            if df.shape[1] == 1 and sep != ",":  # mauvais séparateur
                continue
            return df
        except Exception:
            continue
    bio.seek(0)
    return pd.read_csv(bio, dtype=str)

def read_excel_header3_from_bytes(data: bytes) -> pd.DataFrame:
    return pd.read_excel(BytesIO(data), header=3)

# ================== CENTRES (codes/villes) ==================
CENTER_CANON = {
    "LNE":"Lille","NGE":"Nancy","SIF":"Saclay","PRO":"Paris",
    "RBA":"Rennes","SAM":"Sophia","BSO":"Bordeaux","GRA":"Grenoble","LYS":"Lyon"
}
CENTER_SYNONYMS = {
    "lne":("LNE","Lille"),"lille":("LNE","Lille"),
    "nge":("NGE","Nancy"),"nancy":("NGE","Nancy"),
    "sif":("SIF","Saclay"),"saclay":("SIF","Saclay"),"idf":("SIF","Saclay"),"iledefrance":("SIF","Saclay"),
    "pro":("PRO","Paris"),"paris":("PRO","Paris"),"rocquencourt":("PRO","Paris"),
    "rba":("RBA","Rennes"),"rennes":("RBA","Rennes"),
    "sam":("SAM","Sophia"),"sophia":("SAM","Sophia"),"sophiaantipolis":("SAM","Sophia"),
    "bso":("BSO","Bordeaux"),"bordeaux":("BSO","Bordeaux"),
    "gra":("GRA","Grenoble"),"grenoble":("GRA","Grenoble"),"uga":("GRA","Grenoble"),
    "lys":("LYS","Lyon"),"lyon":("LYS","Lyon"),
}

def detect_center_from_text(text: str) -> Tuple[Optional[str], Optional[str]]:
    t = norm(text)
    for code, ville in CENTER_CANON.items():
        if norm(code) in t: return code, ville
    for key, (code, ville) in CENTER_SYNONYMS.items():
        if key in t: return code, ville
    return None, None

def normalize_center_val(s: str) -> str:
    code, ville = detect_center_from_text(str(s))
    if code: return f"{code} - {ville}"
    return str(s) if s else "AUTRES - Autres"

# ========== PORTFEUILLE (colonnes canons pour l’Excel final) ==========
CANON = [
    "Nom du logiciel","Référence BIL","Référence contrat (legisway)","Lien Legisway",
    "Centres Inria impliqués","Type de licence logiciel",
    "Valorisation","Description (BIL)","Date de dépôt APP","IDDN",
    "Équipe","Auteurs et parts","Logo ?","Commentaires"
]
ALIASES: Dict[str, List[str]] = {
    "Nom du logiciel":["Nom du logiciel","Logiciel","Libellé","Famille","Famille de brevet","Nom"],
    "Référence BIL":["Référence BIL","Ref BIL","BIL","Référence"],
    "Référence contrat (legisway)":["Référence contrat (legisway)","Référence contrat","Ref contrat","Contrat","N° contrat","Numéro de contrat","Contract number","Contract ID","ID contrat"],
    "Lien Legisway":["Lien Legisway","Lien","URL contrat","Hyperlien Legisway"],
    "Centres Inria impliqués":["Centres Inria impliqués","Centre Inria","Centre","Critère 3"],
    "Type de licence logiciel":["Type de licence logiciel","Licence","Type licence","Mots clés"],
    "Valorisation":["Valorisation (licence, cession, projet, dormant, consortium)","Valorisation"],
    "Description (BIL)":["Description (BIL)","Description","Desc BIL"],
    "Date de dépôt APP":["Date de dépôt APP","Date de dépôt","Date depot","Date APP"],
    "IDDN":["IDDN","Num IDDN"],
    "Équipe":["Équipe","Equipe du projet","Équipe du projet","Equipe"],
    "Auteurs et parts":["Auteurs et leurs parts","Auteurs","Auteurs et parts"],
    "Logo ?":["Est-ce qu'il y a un logo ?","Logo ?","Logo"],
    "Commentaires":["Commentaires","Notes","Remarques"],
}

def unify_to_canon(df: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame()
    for k in CANON:
        col = find_col(df, ALIASES.get(k,[k]))
        out[k] = df[col] if col else ""
    out["__nom_lower"] = out["Nom du logiciel"].astype(str).str.strip().str.lower()
    out["__bil_num"] = out["Référence BIL"].astype(str).str.extract(r"(\d+)")
    out["Centres Inria impliqués"] = out["Centres Inria impliqués"].apply(normalize_center_val)
    return out

# ========== LEGISWAY (numéro + lien + match par mots) ==========
LEGIS_FIRST_COL_CAND = ["Contrat","Contract","Titre","Nom","Intitulé"]
LEGIS_NAME_ALIASES = ["Nom du logiciel","Logiciel","Famille","Libellé","Nom"]

def load_legis_with_hyperlinks(file) -> Optional[pd.DataFrame]:
    if file is None: return None
    name, data = get_bytes(file)
    if not data: return None
    wb = load_workbook(BytesIO(data), data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value if cell.value is not None else f"col_{i+1}" for i, cell in enumerate(ws[1])]
    rows = []
    for row in ws.iter_rows(min_row=2):
        row_dict = {}
        for i, cell in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i+1}"
            row_dict[key] = cell.value
        url = None
        if row and row[0] is not None and row[0].hyperlink is not None:
            try: url = row[0].hyperlink.target
            except Exception: url = str(row[0].hyperlink)
        row_dict["__Lien première colonne"] = url
        rows.append(row_dict)
    df = pd.DataFrame(rows)
    try:
        df_pd = pd.read_excel(BytesIO(data), sheet_name=0, dtype=str)
        for col in df_pd.columns:
            if col not in df: df[col] = df_pd[col]
    except Exception:
        pass
    first_col = df.columns[0] if len(df.columns) else "Contrat"
    for cand in LEGIS_FIRST_COL_CAND:
        if cand in df.columns:
            first_col = cand; break
    def extract_num(s): 
        s = "" if s is None else str(s)
        m = re.match(r"\s*([0-9]{4}-\d+)", s)
        return m.group(1) if m else None
    def extract_title(s):
        s = "" if s is None else str(s)
        parts = s.split(" - ",1)
        return parts[1] if len(parts)>1 else s
    df["__Numéro contrat"] = df[first_col].apply(extract_num)
    df["__Intitulé (après numéro)"] = df[first_col].apply(extract_title)
    name_col = find_col(df, LEGIS_NAME_ALIASES)
    df["__legis_name_lower"] = df[name_col].astype(str).str.strip().str.lower() if name_col else ""
    df["__first_col_name"] = first_col
    return df

def legis_find_by_portfolio_ref(ref: str, legis_df: Optional[pd.DataFrame]) -> Tuple[Optional[str], Optional[str]]:
    if legis_df is None or not ref: return None, None
    m = re.search(r"([0-9]{4}-\d+)", str(ref) or "")
    if not m: return None, None
    num = m.group(1)
    sub = legis_df[legis_df["__Numéro contrat"] == num]
    if sub.empty: return None, None
    link = sub["__Lien première colonne"].iloc[0] if "__Lien première colonne" in sub.columns else None
    return num, link

def _tokens(text: str) -> List[str]:
    return [t.lower() for t in re.findall(r"[A-Za-z0-9]+", text or "") if len(t)>=3]

def legis_guess_by_words(soft_name: str, legis_df: Optional[pd.DataFrame]) -> Tuple[Optional[str], Optional[str]]:
    if legis_df is None or not soft_name: return None, None
    want = set(_tokens(soft_name))
    if not want: return None, None
    def score_row(row):
        txt = f"{row.get('__Intitulé (après numéro)','')} {row.get(row.get('__first_col_name',''), '')}"
        have = set(_tokens(str(txt)))
        return len(want & have)
    df = legis_df.copy()
    df["__score"] = df.apply(score_row, axis=1)
    if df.empty or df["__score"].max()==0: return None, None
    top = df.sort_values("__score", ascending=False).iloc[0]
    return top.get("__Numéro contrat"), top.get("__Lien première colonne")

# ========== BIL (lookup + MAJ) ==========
BIL_NAME_ALIASES = ["Logiciel","Nom du logiciel","Libellé","Nom"]
BIL_DATE_ALIASES = ["Date de dépôt","Date depot","Date_dépôt","Date APP"]
BIL_IDDN_ALIASES = ["Num IDDN","IDDN"]
BIL_CENTRE_ALIASES = ["Centre déposant","Centre deposant","Centre","Centre Inria"]

def build_bil_lookup(df_bil: Optional[pd.DataFrame]) -> Tuple[Optional[pd.DataFrame], Optional[str], Optional[str], Optional[str]]:
    if df_bil is None: return None, None, None, None
    name_col = find_col(df_bil, BIL_NAME_ALIASES)
    date_col = find_col(df_bil, BIL_DATE_ALIASES)
    iddn_col = find_col(df_bil, BIL_IDDN_ALIASES)
    centre_col = find_col(df_bil, BIL_CENTRE_ALIASES)
    df_bil["__bil_name_lower"] = df_bil[name_col].astype(str).str.strip().str.lower() if name_col else ""
    return df_bil, date_col, iddn_col, centre_col

def compute_mise_a_jour(row: pd.Series, bil_df: Optional[pd.DataFrame], bil_date_col: Optional[str], bil_iddn_col: Optional[str], bil_centre_col: Optional[str]) -> Tuple[bool, str]:
    if bil_df is None: return False, ""
    name = str(row.get("Nom du logiciel","")).strip().lower()
    if not name: return False, ""
    sub = bil_df[bil_df["__bil_name_lower"] == name]
    if sub.empty: return False, ""
    reasons = []
    port_date = str(row.get("Date de dépôt APP","")).strip()
    bil_date = str(sub.iloc[0][bil_date_col]).strip() if bil_date_col else ""
    if bil_date and port_date and port_date != bil_date: reasons.append("Date de dépôt différente")
    elif bil_date and not port_date: reasons.append("Date de dépôt manquante (PI)")
    port_iddn = str(row.get("IDDN","")).strip()
    bil_iddn = str(sub.iloc[0][bil_iddn_col]).strip() if bil_iddn_col else ""
    if bil_iddn and port_iddn and norm(port_iddn)!=norm(bil_iddn): reasons.append("IDDN différent")
    elif bil_iddn and not port_iddn: reasons.append("IDDN manquant (PI)")
    port_centre = str(row.get("Centres Inria impliqués","")).strip()
    bil_centre = str(sub.iloc[0][bil_centre_col]).strip() if bil_centre_col else ""
    if bil_centre and port_centre and norm(port_centre)!=norm(bil_centre): reasons.append("Centre différent")
    elif bil_centre and not port_centre: reasons.append("Centre manquant (PI)")
    return (len(reasons)>0, ", ".join(reasons))

# ========== EXCEL 1 classeur / 1 feuille par centre ==========
def build_excel_by_center(df_valo: pd.DataFrame) -> BytesIO:
    wb = Workbook(); wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alt_fill = PatternFill("solid", fgColor="F2F2F2")

    df_valo = df_valo.copy()
    df_valo["Centres Inria impliqués"] = df_valo["Centres Inria impliqués"].apply(normalize_center_val)
    parts = df_valo["Centres Inria impliqués"].fillna("AUTRES - Autres").str.split(" - ", n=1, expand=True)
    df_valo["__centre_code"] = parts[0]; df_valo["__centre_ville"] = parts[1].fillna("Autres")

    headers = CANON
    for (code, ville), sub in df_valo.groupby(["__centre_code","__centre_ville"], dropna=False):
        ws = wb.create_sheet(title=f"{code} - {ville}"[:31])
        ws.append(headers)
        for c in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
            ws.column_dimensions[get_column_letter(c)].width = 28
        for i, (_, r) in enumerate(sub.iterrows(), start=2):
            ws.append([r.get(h,"") for h in headers])
            if i % 2 == 0:
                for c in range(1, len(headers)+1):
                    ws.cell(row=i, column=c).fill = alt_fill
        ws.freeze_panes = "A2"

    out = BytesIO(); wb.save(out); out.seek(0); return out

# ========== EXPORTS PI-Planner ==========
def compute_c3(row: pd.Series) -> str:
    _, ville = detect_center_from_text(str(row.get("Centres Inria impliqués","")))
    return ville or ""

def compute_c1_c2(row: pd.Series, bastri_df: Optional[pd.DataFrame]) -> Tuple[str,str]:
    if bastri_df is None or bastri_df.empty: return "",""
    name_col = find_col(bastri_df, ["Équipe","Equipe"])
    c1_col = find_col(bastri_df, ["Critère 1","C1"])
    c2_col = find_col(bastri_df, ["Critère 2","C2"])
    equipe = str(row.get("Équipe","")).strip().lower()
    if not (name_col and equipe): return "",""
    sub = bastri_df[bastri_df[name_col].astype(str).str.strip().str.lower()==equipe]
    if sub.empty: return "",""
    return (str(sub.iloc[0][c1_col]) if c1_col else "",
            str(sub.iloc[0][c2_col]) if c2_col else "")

def build_pi_export(df_src: pd.DataFrame, bastri_df: Optional[pd.DataFrame]) -> pd.DataFrame:
    out = pd.DataFrame()
    out["Famille"] = df_src["Nom du logiciel"]
    out["Titre interne"] = ""
    out["Déposant"] = df_src.apply(lambda r: f"Inria {compute_c3(r)}" if compute_c3(r) else "Inria", axis=1)
    out["Titulaire(s)"] = "Inria"
    out["Mandataire"] = "Inria"
    c1, c2 = [], []
    for _, r in df_src.iterrows():
        a,b = compute_c1_c2(r, bastri_df); c1.append(a); c2.append(b)
    out["Critère 1"] = c1; out["Critère 2"] = c2
    out["Critère 3"] = df_src.apply(compute_c3, axis=1)
    out["Auteur(s)"] = ""; out["Responsable"] = ""
    out["Valorisation"] = df_src.get("Valorisation","")
    out["Dépositaire"] = ""
    out["Date de dépôt"] = df_src.get("Date de dépôt APP","")
    if "Référence BIL" in df_src.columns: out["Référence BIL"] = df_src["Référence BIL"]
    if "Référence contrat (legisway)" in df_src.columns: out["Référence contrat (legisway)"] = df_src["Référence contrat (legisway)"]
    return out

def build_import_from_template(df_src: pd.DataFrame, template_df: Optional[pd.DataFrame], bastri_df: Optional[pd.DataFrame]) -> pd.DataFrame:
    if template_df is None or template_df.empty:
        return build_pi_export(df_src, bastri_df)
    base = build_pi_export(df_src, bastri_df)
    out = pd.DataFrame({c:"" for c in template_df.columns})
    mapping = {
        "Famille":"Famille","Titre interne":"Titre interne","Déposant":"Déposant","Titulaire(s)":"Titulaire(s)",
        "Mandataire":"Mandataire","Critère 1":"Critère 1","Critère 2":"Critère 2","Critère 3":"Critère 3",
        "Auteur(s)":"Auteur(s)","Responsable":"Responsable","Valorisation":"Valorisation",
        "Dépositaire":"Dépositaire","Date de dépôt":"Date de dépôt","Référence BIL":"Référence BIL",
        "Référence contrat (legisway)":"Référence contrat (legisway)"
    }
    for dst, src in mapping.items():
        if dst in out.columns and src in base.columns:
            out[dst] = base[src]
    return out

# ================== UI UPLOADS ==================
st.subheader("1) Fichiers")
file_ip = st.file_uploader("Export PI-Planner (obligatoire)", type=["csv"])
file_portefeuille = st.file_uploader("Portefeuille multi-centres (fichier Excel - trouvable sur MyBox)", type=["xlsx"])
file_bil = st.file_uploader("Export BIL (CSV)", type=["csv"])
file_legis = st.file_uploader("Export Legisway (CSV/Excel)", type=["csv","xlsx","xls"])
file_bastri = st.file_uploader("Export BASTRI (optionnel)", type=["csv","xlsx","xls"])
file_import_tpl = st.file_uploader("Modèle d'import PI-Planner (optionnel - si MAJ)", type=["csv"])

with st.expander("⚙️ Paramètres valorisation"):
    kw_input = st.text_input("Mot-clé de licence = VALORISABLE", value="licence propriétaire")
    license_keywords = [k.strip() for k in kw_input.split(",") if k.strip()]
    st.caption("Valorisable = (n° de contrat Legisway détecté) OU (Type de licence contient l’un des mots-clés)")

st.markdown("---")

# ================== PIPELINE (PI-PLANNER COMME BASE) ==================
if st.button("🚀 En route"):
    if not file_ip:
        st.error("Merci de charger l’export PI-Planner (obligatoire)."); st.stop()

    # 0) Charger PI-Planner (BASE)
    df_ip_raw = parse_csv_or_excel(file_ip)
    if df_ip_raw is None or df_ip_raw.empty:
        st.error("PI-Planner est vide/invalide."); st.stop()

    # Remonter le PI dans le format canon (autant que possible)
    ip_canon = unify_to_canon(df_ip_raw)
    # BIL depuis la colonne Référence (si présente dans l’export PI d’origine)
    ref_col = find_col(df_ip_raw, ["Référence","Reference"])
    if ref_col is not None:
        ip_canon["__bil_num"] = df_ip_raw[ref_col].astype(str).str.extract(r"(\d+)")
        ip_canon["Référence BIL"] = ip_canon["Référence BIL"].replace("", ip_canon["__bil_num"])
    # C3 depuis PI si présent
    c3_col = find_col(df_ip_raw, ["Critère 3","C3"])
    if c3_col:
        ip_canon["Centres Inria impliqués"] = df_ip_raw[c3_col].astype(str).apply(normalize_center_val)

    # 1) Charger Legisway (pour compléter PI avec n° contrat + lien)
    legis_df = load_legis_with_hyperlinks(file_legis)
    # complétion n° + lien sur PI
    for idx, row in ip_canon.iterrows():
        ref = str(row.get("Référence contrat (legisway)",""))
        num, link = legis_find_by_portfolio_ref(ref, legis_df)
        if not num:
            num, link = legis_guess_by_words(str(row.get("Nom du logiciel","")), legis_df)
        if num:
            ip_canon.at[idx,"Référence contrat (legisway)"] = num
            ip_canon.at[idx,"Lien Legisway"] = link or ""

    # 2) Charger BIL pour comparer (MAJ)
    bil_df_raw = parse_csv_or_excel(file_bil) if file_bil else None
    bil_df, bil_date_col, bil_iddn_col, bil_centre_col = build_bil_lookup(bil_df_raw)

    # 3) Charger le Portefeuille multi-feuilles (source de “Nouveaux” + champs complémentaires)
    nouveaux = pd.DataFrame(columns=CANON + ["__nom_lower","__bil_num"])
    if file_portefeuille is not None:
        fname, fbytes = get_bytes(file_portefeuille)
        try:
            xls = pd.ExcelFile(BytesIO(fbytes))
            unified = []
            for sheet in xls.sheet_names:
                df_sheet = pd.read_excel(BytesIO(fbytes), sheet_name=sheet, header=3)
                u = unify_to_canon(df_sheet); unified.append(u)
            pf_all = pd.concat(unified, ignore_index=True) if unified else pd.DataFrame()
        except Exception:
            df_sheet = pd.read_excel(BytesIO(fbytes), header=3)
            pf_all = unify_to_canon(df_sheet)

        # Déterminer ce qui N'EST PAS dans PI-Planner (match par BIL prioritaire, sinon nom)
        ip_bils = set(ip_canon["__bil_num"].dropna().astype(str))
        ip_names = set(ip_canon["__nom_lower"].dropna().astype(str))
        is_new = []
        for _, r in pf_all.iterrows():
            bil = str(r.get("__bil_num","") or "")
            name = str(r.get("__nom_lower","") or "")
            if (bil and bil not in ip_bils) or (not bil and name and name not in ip_names):
                is_new.append(True)
            else:
                is_new.append(False)
        pf_all["__is_new_vs_ip"] = is_new
        nouveaux = pf_all[pf_all["__is_new_vs_ip"]].copy()

    # 4) Flags VALORISABLE (PI & Nouveaux)
    def has_kw(row):
        v = str(row.get("Type de licence logiciel","")).lower().strip()
        if not v: return False
        return any(kw.lower() in v for kw in license_keywords)

    for df_ in (ip_canon, nouveaux):
        if df_.empty: continue
        df_["__has_contract"] = df_["Référence contrat (legisway)"].astype(str).str.contains(r"[0-9]{4}-\d+", na=False)
        df_["__has_kw"] = df_.apply(has_kw, axis=1)
        df_["Valorisable"] = df_["__has_contract"] | df_["__has_kw"]

    # 5) MISE À JOUR (écarts PI vs BIL)
    if bil_df is not None and not ip_canon.empty:
        flags, reasons = [], []
        for _, r in ip_canon.iterrows():
            f, why = compute_mise_a_jour(r, bil_df, bil_date_col, bil_iddn_col, bil_centre_col)
            flags.append(f); reasons.append(why)
        ip_canon["Mise à jour"] = flags; ip_canon["Raison MAJ"] = reasons
    else:
        ip_canon["Mise à jour"] = False; ip_canon["Raison MAJ"] = ""

    # ----------- AFFICHAGE (PI comme base) -----------
    st.subheader("2) Vue PI-Planner (BASE)")
    st.dataframe(ip_canon[CANON + ["Valorisable","Mise à jour","Raison MAJ"]], use_container_width=True)

    st.subheader("3) Nouveaux détectés (absents de PI-Planner)")
    st.dataframe(nouveaux[CANON + ["Valorisable"]], use_container_width=True)

    st.subheader("4) Mises à jour nécessaires (dans PI-Planner)")
    st.dataframe(ip_canon[ip_canon["Mise à jour"]][CANON + ["Raison MAJ"]], use_container_width=True)

    # ----------- SÉLECTION (liberté totale) -----------
    st.markdown("---")
    st.subheader("5) Choix des logiciels à exporter")
    options = (pd.concat([ip_canon[["Nom du logiciel","Valorisable"]],
                          nouveaux[["Nom du logiciel","Valorisable"]]])
               if not nouveaux.empty else ip_canon[["Nom du logiciel","Valorisable"]]).dropna()
    default_sel = options[options["Valorisable"]]["Nom du logiciel"].unique().tolist()
    selected = st.multiselect("Sélectionne pour construire l’Excel et les exports",
                              options=options["Nom du logiciel"].unique().tolist(),
                              default=default_sel)
    selected_df = pd.concat([
        ip_canon[ip_canon["Nom du logiciel"].astype(str).isin(selected)],
        nouveaux[nouveaux["Nom du logiciel"].astype(str).isin(selected)]
    ], ignore_index=True).drop_duplicates(subset=["Nom du logiciel"])

    # ----------- EXCEL FINAL (1 fichier, 1 feuille/centre) -----------
    st.subheader("6) Excel consolidé (1 feuille par centre)")
    excel_bytes = build_excel_by_center(selected_df)
    today = datetime.now(ZoneInfo("Europe/Paris")).strftime("%d-%m-%Y")
    st.download_button("📥 Télécharger l’Excel consolidé",
        data=excel_bytes,
        file_name=f"{today}-portefeuille-centres.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ----------- EXPORTS PI-PLANNER -----------
    bastri_df = parse_csv_or_excel(file_bastri)
    tpl_df = parse_csv_or_excel(file_import_tpl)

    st.subheader("7) Export PI-Planner (enrichi C1/C2/C3)")
    pi_export = build_pi_export(selected_df, bastri_df)
    st.dataframe(pi_export.head(20), use_container_width=True)
    st.download_button("📥 Télécharger l’export PI-Planner (CSV)",
        data=pi_export.to_csv(index=False, sep=";", encoding="utf-8").encode("utf-8"),
        file_name="export_pi_planner.csv",
        mime="text/csv")

    st.subheader("8) Import PI-Planner (aligné sur ton modèle si fourni)")
    import_df = build_import_from_template(selected_df, tpl_df, bastri_df)
    st.dataframe(import_df.head(20), use_container_width=True)
    st.download_button("📥 Télécharger l’import PI-Planner (CSV)",
        data=import_df.to_csv(index=False, sep=";", encoding="utf-8").encode("utf-8"),
        file_name="import_pi_planner.csv",
        mime="text/csv")

    with st.expander("🔧 Debug"):
        if df_ip_raw is not None: st.write("PI-Planner colonnes:", list(df_ip_raw.columns))
        if bil_df_raw is not None: st.write("BIL colonnes:", list(bil_df_raw.columns))
        if legis_df is not None: st.write("Legisway colonnes:", list(legis_df.columns))
        st.write("Alias canon:", ALIASES)