import streamlit as st
import pandas as pd
import re
from typing import Optional, List, Dict, Tuple, Any
from io import BytesIO
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import itertools
import os

# =========================================
#            EN-TÊTE + LOGO
# =========================================
st.set_page_config(page_title="Portefeuille Logiciels – Consolidation & PI-Planner", layout="wide")

LOGO_PATH = "/mnt/data/Logo Agence Créative Minimaliste Moderne Blanc et Bleu .png"
c1, c2 = st.columns([1, 5])
with c1:
    if os.path.exists(LOGO_PATH):
        st.image(LOGO_PATH, caption="", use_column_width=True)
with c2:
    st.title("🗂️ Portefeuille multi-centres → Consolidation & Export PI-Planner")

# =========================================
#                HELPERS
# =========================================
def norm(s: Any) -> str:
    """Minuscule + supprime espaces/ponctuation (pour comparaisons robustes)."""
    if s is None:
        return ""
    x = str(s).strip().lower()
    x = re.sub(r"[^a-z0-9]+", "", x)
    return x

def find_col(df: pd.DataFrame, aliases: List[str]) -> Optional[str]:
    """Trouve une colonne en testant une liste d'alias (normalisés)."""
    norm_cols = {norm(c): c for c in df.columns}
    for alias in aliases:
        a = norm(alias)
        if a in norm_cols:
            return norm_cols[a]
    return None

def get_bytes(file) -> Tuple[str, bytes]:
    """Retourne (nom, bytes) d'un UploadedFile (ou file-like)."""
    if file is None:
        return "", b""
    name = getattr(file, "name", "uploaded")
    try:
        file.seek(0)
    except Exception:
        pass
    data = file.read()
    return name, data

def parse_csv_or_excel(file) -> Optional[pd.DataFrame]:
    """Charge CSV (essaye ; puis , puis tab) ou Excel."""
    if file is None:
        return None
    name, data = get_bytes(file)
    if not data:
        return None
    bio = BytesIO(data)
    if name.lower().endswith((".xlsx", ".xls")):
        return pd.read_excel(bio)
    # CSV
    for sep in [";", ",", "\t"]:
        try:
            bio.seek(0)
            df = pd.read_csv(bio, sep=sep, engine="python", dtype=str)
            if df.shape[1] == 1 and sep != ",":
                continue
            return df
        except Exception:
            continue
    bio.seek(0)
    return pd.read_csv(bio, dtype=str)

def read_excel_header3_from_bytes(data: bytes) -> pd.DataFrame:
    """Lit un Excel en mémoire avec header=3 (4e ligne = en-têtes)."""
    return pd.read_excel(BytesIO(data), header=3)

# =========================================
#                CENTRES
# =========================================
CENTER_CANON = {
    "LNE": "Lille", "NGE": "Nancy", "SIF": "Saclay", "PRO": "Paris",
    "RBA": "Rennes", "SAM": "Sophia", "BSO": "Bordeaux", "GRA": "Grenoble", "LYS": "Lyon",
}
CENTER_SYNONYMS = {
    "lne": ("LNE", "Lille"), "lille": ("LNE", "Lille"),
    "nge": ("NGE", "Nancy"), "nancy": ("NGE", "Nancy"),
    "sif": ("SIF", "Saclay"), "saclay": ("SIF", "Saclay"), "idf": ("SIF", "Saclay"), "iledefrance": ("SIF", "Saclay"),
    "pro": ("PRO", "Paris"), "paris": ("PRO", "Paris"), "rocquencourt": ("PRO", "Paris"),
    "rba": ("RBA", "Rennes"), "rennes": ("RBA", "Rennes"),
    "sam": ("SAM", "Sophia"), "sophia": ("SAM", "Sophia"), "sophiaantipolis": ("SAM", "Sophia"),
    "bso": ("BSO", "Bordeaux"), "bordeaux": ("BSO", "Bordeaux"),
    "gra": ("GRA", "Grenoble"), "grenoble": ("GRA", "Grenoble"), "uga": ("GRA", "Grenoble"),
    "lys": ("LYS", "Lyon"), "lyon": ("LYS", "Lyon"),
}

def detect_center_from_text(text: str) -> Tuple[Optional[str], Optional[str]]:
    t = norm(text)
    for code, ville in CENTER_CANON.items():
        if norm(code) in t:
            return code, ville
    for key, (code, ville) in CENTER_SYNONYMS.items():
        if key in t:
            return code, ville
    return None, None

def auto_detect_file_center(file_name: str, file_bytes: bytes) -> Tuple[str, str]:
    code, ville = detect_center_from_text(file_name)
    if code:
        return code, ville
    try:
        xls = pd.ExcelFile(BytesIO(file_bytes))
        for sh in xls.sheet_names:
            code, ville = detect_center_from_text(sh)
            if code:
                return code, ville
        for sh in xls.sheet_names:
            head = pd.read_excel(BytesIO(file_bytes), sheet_name=sh, header=None, nrows=5)
            flat = " ".join([str(v) for v in itertools.chain.from_iterable(head.values.tolist()) if pd.notna(v)])
            code, ville = detect_center_from_text(flat)
            if code:
                return code, ville
    except Exception:
        pass
    return "AUTRES", "Autres"

# =========================================
#        COLONNES CANON (portefeuille)
# =========================================
CANON = [
    "Nom du logiciel", "Référence BIL", "Référence contrat (legisway)", "Lien Legisway",
    "Centres Inria impliqués", "Type de licence logiciel",
    "Valorisation", "Description (BIL)", "Date de dépôt APP", "IDDN",
    "Équipe", "Auteurs et parts", "Logo ?", "Commentaires"
]
ALIASES: Dict[str, List[str]] = {
    "Nom du logiciel": ["Nom du logiciel", "Logiciel", "Libellé", "Nom"],
    "Référence BIL": ["Référence BIL", "Ref BIL", "BIL", "Réf BIL"],
    "Référence contrat (legisway)": ["Référence contrat (legisway)", "Référence contrat", "Reference contrat", "Ref contrat", "Contrat", "Legisway", "N° contrat", "Numéro de contrat", "Numero de contrat", "Contract number", "Contract ID", "ID contrat"],
    "Lien Legisway": ["Lien Legisway", "Lien", "URL contrat", "Hyperlien Legisway"],
    "Centres Inria impliqués": ["Centres Inria impliqués", "Centre Inria", "Centre", "Centre déposant", "Centre deposant"],
    "Type de licence logiciel": ["Type de licence logiciel", "Licence", "Type licence", "Mots clés"],
    "Valorisation": ["Valorisation (licence, cession, projet, dormant, consortium)", "Valorisation"],
    "Description (BIL)": ["Description (BIL)", "Description", "Desc BIL"],
    "Date de dépôt APP": ["Date de dépôt APP", "Date de dépôt", "Date depot", "Date APP"],
    "IDDN": ["IDDN", "Num IDDN"],
    "Équipe": ["Equipe du projet", "Équipe du projet", "Équipe", "Equipe"],
    "Auteurs et parts": ["Auteurs et leurs parts", "Auteurs", "Auteurs et parts"],
    "Logo ?": ["Est-ce qu'il y a un logo ?", "Logo ?", "Logo"],
    "Commentaires": ["Commentaires", "Notes", "Remarques"]
}

def unify_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame()
    for canon in CANON:
        col = find_col(df, ALIASES.get(canon, [canon]))
        out[canon] = df[col] if col else ""
    out["__nom_lower"] = out["Nom du logiciel"].astype(str).str.strip().str.lower()
    out["__bil_num"] = out["Référence BIL"].astype(str).str.extract(r"(\d+)")
    return out

# =========================================
#             LEGISWAY (++ hyperliens)
# =========================================
LEGIS_NUM_ALIASES = [
    "Numéro de contrat", "Numero de contrat", "N° contrat", "Contract number",
    "Contract ID", "ID contrat", "Référence contrat", "Reference contrat", "Réf contrat", "Ref contrat"
]
LEGIS_NAME_ALIASES = ["Nom du logiciel", "Logiciel", "Famille", "Libellé", "Nom"]
LEGIS_FIRST_COL_CAND = ["Contrat", "Contract", "Titre", "Nom", "Intitulé"]

def load_legis_with_hyperlinks(file) -> Optional[pd.DataFrame]:
    if file is None:
        return None
    name, data = get_bytes(file)
    if not data:
        return None
    # openpyxl pour hyperliens
    wb = load_workbook(BytesIO(data), data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value if cell.value is not None else f"col_{i+1}" for i, cell in enumerate(ws[1])]
    rows = []
    for row in ws.iter_rows(min_row=2):
        row_dict = {}
        for i, cell in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i+1}"
            row_dict[key] = cell.value
        url = None
        if row and row[0] is not None and row[0].hyperlink is not None:
            try:
                url = row[0].hyperlink.target
            except Exception:
                url = str(row[0].hyperlink)
        row_dict["__Lien première colonne"] = url
        rows.append(row_dict)
    df = pd.DataFrame(rows)
    # pandas pour robustesse des types
    try:
        df_pd = pd.read_excel(BytesIO(data), sheet_name=0, dtype=str)
        for col in df_pd.columns:
            if col not in df.columns:
                df[col] = df_pd[col]
    except Exception:
        pass
    # première colonne "descriptive"
    first_col = df.columns[0] if len(df.columns) else "Contrat"
    for cand in LEGIS_FIRST_COL_CAND:
        if cand in df.columns:
            first_col = cand
            break
    # extractions
    def extract_contract_num(text: Any) -> Optional[str]:
        s = "" if text is None else str(text)
        m = re.match(r"\s*([0-9]{4}-\d+)", s)
        return m.group(1) if m else None
    def extract_title(text: Any) -> str:
        s = "" if text is None else str(text)
        parts = s.split(" - ", 1)
        return parts[1] if len(parts) > 1 else s
    df["__Numéro contrat"] = df[first_col].apply(extract_contract_num)
    df["__Intitulé (après numéro)"] = df[first_col].apply(extract_title)
    df["__First col name"] = first_col
    # Index nom logiciel (pour match par mots)
    name_col = find_col(df, LEGIS_NAME_ALIASES)
    if name_col:
        df["__legis_name_lower"] = df[name_col].astype(str).str.strip().str.lower()
    else:
        df["__legis_name_lower"] = ""
    return df

def legis_find_number_by_contract_ref(portfolio_ref: str, legis_df: Optional[pd.DataFrame]) -> Tuple[Optional[str], Optional[str]]:
    """Retourne (numéro, lien) si portfolio_ref matche un numéro Legisway."""
    if legis_df is None or not portfolio_ref:
        return None, None
    ref_num = re.search(r"([0-9]{4}-\d+)", str(portfolio_ref) or "")
    if not ref_num:
        return None, None
    num = ref_num.group(1)
    match = legis_df[legis_df["__Numéro contrat"] == num]
    if match.empty:
        return None, None
    link = match["__Lien première colonne"].iloc[0] if "__Lien première colonne" in match.columns else None
    return num, link

def tokens(text: str) -> List[str]:
    """Petits tokens >=3 caractères pour match par mots."""
    toks = re.findall(r"[A-Za-z0-9]+", (text or ""))
    return [t.lower() for t in toks if len(t) >= 3]

def legis_guess_number_by_words(soft_name: str, legis_df: Optional[pd.DataFrame]) -> Tuple[Optional[str], Optional[str]]:
    """Si pas de numéro dans le portefeuille, essaie de retrouver un contrat Legisway
       dont la 1re colonne contient suffisamment de mots du nom de logiciel."""
    if legis_df is None or not soft_name:
        return None, None
    name_toks = set(tokens(soft_name))
    if not name_toks:
        return None, None
    # score = taille de l'intersection avec "__Intitulé (après numéro)"
    def score_row(row) -> int:
        text = str(row.get("__Intitulé (après numéro)", "")) + " " + str(row.get(row.get("__First col name", ""), ""))
        rtoks = set(tokens(text))
        return len(name_toks & rtoks)
    df = legis_df.copy()
    df["__score"] = df.apply(score_row, axis=1)
    df = df.sort_values("__score", ascending=False)
    top = df.iloc[0] if not df.empty else None
    if top is None or top["__score"] == 0:
        return None, None
    return top.get("__Numéro contrat"), top.get("__Lien première colonne")

# =========================================
#            BIL (lookup & MAJ)
# =========================================
BIL_NAME_ALIASES = ["Logiciel", "Nom du logiciel", "Libellé", "Nom"]
BIL_DATE_ALIASES = ["Date de dépôt", "Date depot", "Date_dépôt", "Date APP"]
BIL_IDDN_ALIASES = ["Num IDDN", "IDDN"]
BIL_CENTRE_ALIASES = ["Centre déposant", "Centre deposant", "Centre", "Centre Inria"]

def build_bil_lookup(df_bil: Optional[pd.DataFrame]) -> Tuple[Optional[pd.DataFrame], Optional[str], Optional[str], Optional[str]]:
    if df_bil is None:
        return None, None, None, None
    name_col = find_col(df_bil, BIL_NAME_ALIASES)
    date_col = find_col(df_bil, BIL_DATE_ALIASES)
    iddn_col = find_col(df_bil, BIL_IDDN_ALIASES)
    centre_col = find_col(df_bil, BIL_CENTRE_ALIASES)
    if name_col:
        df_bil["__bil_name_lower"] = df_bil[name_col].astype(str).str.strip().str.lower()
    else:
        df_bil["__bil_name_lower"] = ""
    return df_bil, date_col, iddn_col, centre_col

def compute_mise_a_jour(row: pd.Series, bil_df: Optional[pd.DataFrame], bil_date_col: Optional[str], bil_iddn_col: Optional[str], bil_centre_col: Optional[str]) -> Tuple[bool, str]:
    if bil_df is None:
        return False, ""
    name = str(row.get("Nom du logiciel", "")).strip().lower()
    if not name:
        return False, ""
    sub = bil_df[bil_df["__bil_name_lower"] == name]
    if sub.empty:
        return False, ""
    reasons = []
    port_date = str(row.get("Date de dépôt APP", "")).strip()
    bil_date = str(sub.iloc[0][bil_date_col]).strip() if bil_date_col else ""
    if bil_date and port_date and port_date != bil_date:
        reasons.append("Date de dépôt différente")
    elif bil_date and not port_date:
        reasons.append("Date de dépôt manquante (portefeuille)")
    port_iddn = str(row.get("IDDN", "")).strip()
    bil_iddn = str(sub.iloc[0][bil_iddn_col]).strip() if bil_iddn_col else ""
    if bil_iddn and port_iddn and norm(port_iddn) != norm(bil_iddn):
        reasons.append("IDDN différent")
    elif bil_iddn and not port_iddn:
        reasons.append("IDDN manquant (portefeuille)")
    port_centre = str(row.get("Centres Inria impliqués", "")).strip()
    bil_centre = str(sub.iloc[0][bil_centre_col]).strip() if bil_centre_col else ""
    if bil_centre and port_centre and norm(port_centre) != norm(bil_centre):
        reasons.append("Centre différent")
    elif bil_centre and not port_centre:
        reasons.append("Centre manquant (portefeuille)")
    return (len(reasons) > 0, ", ".join(reasons))

# =========================================
#           EXCEL (1 feuille / centre)
# =========================================
def build_excel_by_center(df_valo: pd.DataFrame) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alt_fill = PatternFill("solid", fgColor="F2F2F2")

    df_valo = df_valo.copy()

    # normalise "Centres Inria impliqués" => "CODE - Ville"
    def normalize_center_val(s):
        code, ville = detect_center_from_text(str(s))
        if code:
            return f"{code} - {ville}"
        return str(s) if s else "AUTRES - Autres"
    df_valo["Centres Inria impliqués"] = df_valo["Centres Inria impliqués"].apply(normalize_center_val)

    centres_split = df_valo["Centres Inria impliqués"].fillna("AUTRES - Autres").str.split(" - ", n=1, expand=True)
    df_valo["__centre_code"] = centres_split[0]
    df_valo["__centre_ville"] = centres_split[1].fillna("Autres")

    headers = CANON  # garde la forme générale + ajoute Lien Legisway
    for (code, ville), sub in df_valo.groupby(["__centre_code", "__centre_ville"], dropna=False):
        sheet_name = f"{code} - {ville}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
            ws.column_dimensions[get_column_letter(c)].width = 28

        for i, (_, r) in enumerate(sub.iterrows(), start=2):
            ws.append([r.get(h, "") for h in headers])
            if i % 2 == 0:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=i, column=c).fill = alt_fill
        ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# =========================================
#        EXPORTS PI-PLANNER (NOUVEAUX)
# =========================================
def compute_c3(row: pd.Series) -> str:
    raw = str(row.get("Centres Inria impliqués", ""))
    _, ville = detect_center_from_text(raw)
    return ville or ""

def compute_c1_c2(row: pd.Series, bastri_df: Optional[pd.DataFrame]) -> Tuple[str, str]:
    # Sans BASTRI fourni, on ne devine pas : précision > devinette
    if bastri_df is None or bastri_df.empty:
        return "", ""
    # exemple simple : si la colonne "Équipe" existe des deux côtés, on transpose vers C1/C2 via mapping fourni.
    # Tu pourras adapter/élargir le mapping ici lorsque tu me donnes la table exacte BASTRI.
    equipe = str(row.get("Équipe", "")).strip().lower()
    if not equipe:
        return "", ""
    # recherche naive dans bastri
    name_col = find_col(bastri_df, ["Équipe", "Equipe"])
    c1_col = find_col(bastri_df, ["Critère 1", "C1"])
    c2_col = find_col(bastri_df, ["Critère 2", "C2"])
    if not name_col:
        return "", ""
    sub = bastri_df[bastri_df[name_col].astype(str).str.strip().str.lower() == equipe]
    if sub.empty:
        return "", ""
    c1 = str(sub.iloc[0][c1_col]) if c1_col else ""
    c2 = str(sub.iloc[0][c2_col]) if c2_col else ""
    return c1, c2

def build_pi_planner_export_from_df(df_src: pd.DataFrame, bastri_df: Optional[pd.DataFrame]) -> pd.DataFrame:
    # Colonnes enrichies (C1, C2, C3)
    out = pd.DataFrame()
    out["Famille"] = df_src["Nom du logiciel"]
    out["Titre interne"] = ""
    out["Déposant"] = df_src.apply(lambda r: f"Inria {compute_c3(r)}" if compute_c3(r) else "Inria", axis=1)
    out["Titulaire(s)"] = "Inria"
    out["Mandataire"] = "Inria"
    # C1 / C2 optionnels via BASTRI
    c1_list, c2_list = [], []
    for _, r in df_src.iterrows():
        c1, c2 = compute_c1_c2(r, bastri_df)
        c1_list.append(c1)
        c2_list.append(c2)
    out["Critère 1"] = c1_list
    out["Critère 2"] = c2_list
    out["Critère 3"] = df_src.apply(compute_c3, axis=1)
    out["Auteur(s)"] = ""
    out["Responsable"] = ""
    out["Valorisation"] = df_src.get("Valorisation", "")
    out["Dépositaire"] = ""
    out["Date de dépôt"] = df_src.get("Date de dépôt APP", "")
    # Ajouts utiles si présents
    if "Référence BIL" in df_src.columns:
        out["Référence BIL"] = df_src["Référence BIL"]
    if "Référence contrat (legisway)" in df_src.columns:
        out["Référence contrat (legisway)"] = df_src["Référence contrat (legisway)"]
    return out

def build_import_from_template(df_src: pd.DataFrame, template_df: Optional[pd.DataFrame]) -> pd.DataFrame:
    """
    Si un modèle d'import (CSV) est fourni, fabriquer un DataFrame avec EXACTEMENT
    les colonnes du modèle, en remplissant tout ce qu'on peut depuis df_src.
    Sinon, on renvoie l'export standard PI-Planner enrichi.
    """
    if template_df is None or template_df.empty:
        return build_pi_planner_export_from_df(df_src, bastri_df=None)

    cols = list(template_df.columns)
    base = build_pi_planner_export_from_df(df_src, bastri_df=None)
    out = pd.DataFrame({c: "" for c in cols})
    # mappe les champs connus
    mapping = {
        "Famille": "Famille",
        "Titre interne": "Titre interne",
        "Déposant": "Déposant",
        "Titulaire(s)": "Titulaire(s)",
        "Mandataire": "Mandataire",
        "Critère 1": "Critère 1",
        "Critère 2": "Critère 2",
        "Critère 3": "Critère 3",
        "Auteur(s)": "Auteur(s)",
        "Responsable": "Responsable",
        "Valorisation": "Valorisation",
        "Dépositaire": "Dépositaire",
        "Date de dépôt": "Date de dépôt",
        "Référence BIL": "Référence BIL",
        "Référence contrat (legisway)": "Référence contrat (legisway)",
    }
    for dst_col, src_col in mapping.items():
        if dst_col in out.columns and src_col in base.columns:
            out[dst_col] = base[src_col]
    return out

# =========================================
#                    UI
# =========================================
st.subheader("1) Charger les fichiers")
portfolios = st.file_uploader("Portefeuilles (Excel) – sélection multiple", type=["xlsx"], accept_multiple_files=True)
file_bil = st.file_uploader("Export BIL (CSV)", type=["csv"])
file_ip = st.file_uploader("Export PI-Planner (FamillesBrevet.csv)", type=["csv"])
file_legis = st.file_uploader("Export Legisway (CSV/Excel)", type=["csv", "xlsx", "xls"])
file_import_tpl = st.file_uploader("Modèle d'import PI-Planner (CSV) – optionnel", type=["csv"])
file_bastri = st.file_uploader("Export BASTRI (Équipe → C1/C2) – optionnel", type=["csv", "xlsx", "xls"])

with st.expander("⚙️ Paramètres"):
    kw_input = st.text_input("Mot-clé de licence à considérer comme VALORISABLE (ex.: licence propriétaire)", value="licence propriétaire")
    license_keywords = [k.strip() for k in kw_input.split(",") if k.strip()]
    st.caption("Règle: Valorisable = (numéro de contrat Legisway détecté) OU (Type de licence contient l’un des mots-clés ci-dessus)")

# ---- UI override centres par fichier ----
st.subheader("1bis) Centres détectés (corrige si besoin)")
override_centers: Dict[str, Tuple[str, str]] = {}
if portfolios:
    for f in portfolios:
        fname, fbytes = get_bytes(f)
        auto_code, auto_city = auto_detect_file_center(fname, fbytes)
        options = [("AUTRES", "Autres")] + list(CENTER_CANON.items())
        labels = [f"{c} – {v}" for c, v in options]
        default_idx = 0
        for i, (c, v) in enumerate(options):
            if c == auto_code:
                default_idx = i
                break
        sel = st.selectbox(
            f"Centre pour « {fname} »",
            options=list(range(len(options))),
            index=default_idx,
            format_func=lambda i: labels[i]
        )
        chosen_code, chosen_city = options[sel]
        override_centers[fname] = (chosen_code, chosen_city)

st.markdown("---")

# =========================================
#           PIPELINE PRINCIPAL
# =========================================
if st.button("🚀 Consolider, filtrer et préparer l'export"):
    # --- Entrées minimales ---
    if not portfolios:
        st.error("Merci de charger au moins un fichier **Portefeuille**.")
        st.stop()
    if not file_ip:
        st.error("Merci de charger l'**export PI-Planner** (FamillesBrevet.csv).")
        st.stop()

    # PI-Planner (pour dédoublonnage si besoin)
    df_ip = parse_csv_or_excel(file_ip)
    col_ip_reference = find_col(df_ip, ["Référence", "Reference"])
    if col_ip_reference is not None:
        df_ip["__ip_bil"] = df_ip[col_ip_reference].astype(str).str.extract(r"(\d+)")
    col_ip_famille = find_col(df_ip, ["Famille de brevet", "Famille", "Libellé", "Nom"])
    existing_bil = set(df_ip["__ip_bil"].dropna().astype(str)) if "__ip_bil" in df_ip.columns else set()
    existing_names = set(df_ip[col_ip_famille].dropna().astype(str).str.strip().str.lower()) if col_ip_famille else set()

    # Legisway
    legis_df = load_legis_with_hyperlinks(file_legis)

    # BIL (pour MAJ + aide centre)
    df_bil_raw = parse_csv_or_excel(file_bil) if file_bil else None
    bil_df, bil_date_col, bil_iddn_col, bil_centre_col = build_bil_lookup(df_bil_raw)

    # BASTRI (optionnel, pour C1/C2)
    bastri_df = parse_csv_or_excel(file_bastri) if file_bastri else None

    # Unification & consolidation des portefeuilles
    unified_all = []
    for f in portfolios:
        fname, fbytes = get_bytes(f)
        try:
            df_raw = read_excel_header3_from_bytes(fbytes)
        except Exception:
            df_raw = pd.read_excel(BytesIO(fbytes))
        uni = unify_columns(df_raw)

        # Injecte/normalise le centre
        center_cell = str(uni["Centres Inria impliqués"].iloc[0]) if len(uni) else ""
        if not center_cell.strip():
            code, ville = override_centers.get(fname, auto_detect_file_center(fname, fbytes))
            uni["Centres Inria impliqués"] = f"{code} - {ville}"
        else:
            code, ville = detect_center_from_text(center_cell)
            if code:
                uni["Centres Inria impliqués"] = f"{code} - {ville}"

        # Legisway: remplir numéro + lien si possible
        # 1) si Réf contrat présente → valider/compléter via Legisway
        ref_val = str(uni["Référence contrat (legisway)"].iloc[0]) if len(uni) else ""
        num, link = legis_find_number_by_contract_ref(ref_val, legis_df)
        if num:
            uni["Référence contrat (legisway)"] = num
            if "Lien Legisway" in uni.columns:
                uni["Lien Legisway"] = link or ""
        # 2) sinon tenter par mots (nom logiciel ↔ 1re colonne)
        else:
            for idx, row in uni.iterrows():
                name = str(row.get("Nom du logiciel", ""))
                guessed_num, guessed_link = legis_guess_number_by_words(name, legis_df)
                if guessed_num:
                    uni.at[idx, "Référence contrat (legisway)"] = guessed_num
                    if "Lien Legisway" in uni.columns:
                        uni.at[idx, "Lien Legisway"] = guessed_link or ""

        unified_all.append(uni)

    df_all = pd.concat(unified_all, ignore_index=True)

    # Flags Valorisable & Mise à jour
    def has_keyword_valo(row: pd.Series, keywords: List[str]) -> bool:
        val = str(row.get("Type de licence logiciel", "")).strip().lower()
        if not val:
            return False
        for kw in keywords:
            if kw.lower() in val:
                return True
        return False

    df_all["__has_contract_number"] = df_all["Référence contrat (legisway)"].astype(str).str.contains(r"[0-9]{4}-\d+", na=False)
    df_all["__has_kw_valo"] = df_all.apply(lambda r: has_keyword_valo(r, license_keywords), axis=1)
    df_all["Valorisable"] = df_all["__has_contract_number"] | df_all["__has_kw_valo"]

    if bil_df is not None:
        maj_flags, maj_reasons = [], []
        for _, row in df_all.iterrows():
            flag, reason = compute_mise_a_jour(row, bil_df, bil_date_col, bil_iddn_col, bil_centre_col)
            maj_flags.append(flag)
            maj_reasons.append(reason)
        df_all["Mise à jour"] = maj_flags
        df_all["Raison MAJ"] = maj_reasons
    else:
        df_all["Mise à jour"] = False
        df_all["Raison MAJ"] = ""

    # Affichage : tous / non valorisables
    st.subheader("2) TOUS les logiciels extraits")
    st.dataframe(df_all[CANON + ["Valorisable", "Mise à jour", "Raison MAJ"]], use_container_width=True)

    st.subheader("3) Logiciels non valorisables (seront exclus)")
    df_non_valo = df_all[~df_all["Valorisable"]]
    st.dataframe(df_non_valo[CANON], use_container_width=True)

    # Sélection manuelle (liberté totale)
    st.subheader("4) Choix des logiciels à exporter")
    default_sel = df_all[df_all["Valorisable"]]["Nom du logiciel"].dropna().astype(str).unique().tolist()
    selected = st.multiselect("Sélectionne les logiciels à inclure dans l’Excel & les exports",
                              options=df_all["Nom du logiciel"].dropna().astype(str).unique().tolist(),
                              default=default_sel)
    df_selected = df_all[df_all["Nom du logiciel"].astype(str).isin(selected)].copy()

    # Excel consolidé (feuilles par centre)
    st.markdown("---")
    st.subheader("5) Générer l’Excel consolidé (1 feuille par centre)")
    excel_bytes = build_excel_by_center(df_selected)
    today_str = datetime.now(ZoneInfo("Europe/Paris")).strftime("%d-%m-%Y")
    excel_name = f"{today_str}-portefeuille-centres.xlsx"
    st.download_button(
        "📥 Télécharger l’Excel consolidé",
        data=excel_bytes,
        file_name=excel_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Export PI-Planner (C1/C2 via BASTRI si fourni, C3 = centre)
    st.subheader("6) Export PI-Planner (enrichi C1/C2/C3)")
    pi_export = build_pi_planner_export_from_df(df_selected, bastri_df=bastri_df)
    st.dataframe(pi_export.head(20), use_container_width=True)
    st.download_button(
        "📥 Télécharger l’export PI-Planner (CSV)",
        pi_export.to_csv(index=False, sep=";", encoding="utf-8").encode("utf-8"),
        file_name="export_pi_planner.csv",
        mime="text/csv"
    )

    # Import PI-Planner aligné sur modèle (si fourni)
    tpl_df = parse_csv_or_excel(file_import_tpl)
    import_df = build_import_from_template(df_selected, template_df=tpl_df)
    st.subheader("7) Fichier d'import PI-Planner (aligné sur le modèle si fourni)")
    st.dataframe(import_df.head(20), use_container_width=True)
    st.download_button(
        "📥 Télécharger l’import PI-Planner (CSV)",
        import_df.to_csv(index=False, sep=";", encoding="utf-8").encode("utf-8"),
        file_name="import_pi_planner.csv",
        mime="text/csv"
    )

    with st.expander("🔧 Debug colonnes détectées"):
        if legis_df is not None:
            st.write("Legisway:", list(legis_df.columns))
        if df_bil_raw is not None:
            st.write("BIL:", list(df_bil_raw.columns))
        if df_ip is not None:
            st.write("PI-Planner export:", list(df_ip.columns))
        st.write("Alias portefeuille:", ALIASES)
